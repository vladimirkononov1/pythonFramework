import openpyxl


class HomePageData:

    test_HomePage_data = [{"name": "Rahul Shetty", "email": "test@test.com", "gender": "Male", "password": "12345"},
                            {"name": "Anshika Shetty", "email": "test@test.com", "gender": "Female", "password": "12345"}]

    @staticmethod
    def getTestData(test_case_name):
        # book = openpyxl.load_workbook("C:\\Users\\vladi\\Documents\\PythonDemo.xlsx")
        book = openpyxl.load_workbook("C:\\Users\\vladi\PythonSelFramework\\PythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
